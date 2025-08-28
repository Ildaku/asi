from datetime import datetime, timedelta
from flask import render_template, redirect, url_for, flash, request, jsonify, send_file
from app import app, db
from app.models import (
    RawMaterial, RecipeTemplate as Recipe, Product, RawMaterialType,
    RecipeItem as RecipeIngredient, ProductionPlan, PlanStatus,
    ProductionBatch, MaterialBatch, BatchMaterial, User, UserRole
)
from app.forms import (
    RawMaterialForm, ProductForm, RecipeForm, RecipeIngredientForm,
    RawMaterialTypeForm, ProductionPlanForm, ProductionBatchForm,
    ProductionStatusForm, BatchIngredientForm, RawMaterialUsageReportForm,
    ProductionStatisticsForm, LoginForm
)
from app.utils import (
    create_excel_report, style_header_row, adjust_column_width,
    save_excel_report, format_datetime
)
from sqlalchemy import func, cast, String
from sqlalchemy.orm import joinedload
from flask_login import login_user, logout_user, login_required, current_user
from app.decorators import admin_required, operator_required
from flask_migrate import upgrade
import subprocess
from openpyxl import Workbook
from io import BytesIO

@app.route('/')
@login_required
def index():
    return render_template('index.html')

# --- Только для ADMIN ---
@app.route('/raw_material_types', methods=['GET', 'POST'])
@admin_required
def raw_material_types():
    form = RawMaterialTypeForm()
    if form.validate_on_submit():
        t = RawMaterialType(name=form.name.data)
        db.session.add(t)
        db.session.commit()
        flash('Вид сырья добавлен!', 'success')
        return redirect(url_for('raw_material_types'))
    types = RawMaterialType.query.all()
    return render_template('raw_material_types.html', form=form, types=types)

@app.route('/raw_material_types/delete/<int:id>', methods=['POST'])
@admin_required
def delete_raw_material_type(id):
    t = RawMaterialType.query.get_or_404(id)
    db.session.delete(t)
    db.session.commit()
    flash('Вид сырья удалён!', 'success')
    return redirect(url_for('raw_material_types'))

# --- Сырьё (партии) ---

def sync_material_batch_numbers(raw_material_id):
    """Синхронизирует номера партий во всех связанных записях material_batches"""
    
    try:
        raw_material = RawMaterial.query.get(raw_material_id)
        if not raw_material:
            return False
        
        # Находим все material_batches с этим сырьём
        material_batches = MaterialBatch.query.filter_by(
            material_id=raw_material_id
        ).all()
        
        updated_count = 0
        for mb in material_batches:
            if mb.batch_number != raw_material.batch_number:
                mb.batch_number = raw_material.batch_number
                updated_count += 1
        
        if updated_count > 0:
            db.session.commit()
            return updated_count
        
        return 0
        
    except Exception as e:
        # В случае ошибки откатываем изменения
        db.session.rollback()
        print(f"Ошибка при синхронизации номеров партий: {e}")
        return False

@app.route('/raw_materials', methods=['GET', 'POST'])
@login_required
def raw_materials():
    form = RawMaterialForm()
    form.type_id.choices = [(t.id, t.name) for t in RawMaterialType.query.all()]
    if form.validate_on_submit():
        # Проверяем права на создание
        if not current_user.is_admin():
            flash('Доступ запрещен. Требуются права администратора для создания.', 'error')
            return redirect(url_for('raw_materials'))
        
        raw = RawMaterial(
            name=form.batch_number.data,
            type_id=form.type_id.data,
            batch_number=form.batch_number.data,
            quantity_kg=form.quantity_kg.data,
            date_received=form.date_received.data or None,
            expiration_date=form.expiration_date.data or None
        )
        db.session.add(raw)
        db.session.commit()
        flash('Партия сырья добавлена!', 'success')
        return redirect(url_for('raw_materials'))
    # Доступное сырьё
    materials = RawMaterial.query.filter(RawMaterial.quantity_kg > 0).order_by(RawMaterial.created_at.desc()).all()
    # Выработанное сырьё
    used_up_materials = RawMaterial.query.filter(RawMaterial.quantity_kg == 0).order_by(RawMaterial.created_at.desc()).all()
    
    # Добавляем информацию о днях до истечения срока годности
    today = datetime.now().date()
    for material in materials:
        if material.expiration_date:
            expiration_date = material.expiration_date.date()
            days_until_expiry = (expiration_date - today).days
            material.days_until_expiry = days_until_expiry
        else:
            material.days_until_expiry = None
    for material in used_up_materials:
        if material.expiration_date:
            expiration_date = material.expiration_date.date()
            days_until_expiry = (expiration_date - today).days
            material.days_until_expiry = days_until_expiry
        else:
            material.days_until_expiry = None
    
    return render_template('raw_materials.html', form=form, materials=materials, used_up_materials=used_up_materials)

@app.route('/raw_materials/edit/<int:id>', methods=['GET', 'POST'])

@admin_required
def edit_raw_material(id):
    material = RawMaterial.query.get_or_404(id)
    form = RawMaterialForm(obj=material)
    form.type_id.choices = [(t.id, t.name) for t in RawMaterialType.query.all()]
    if form.validate_on_submit():
        # Сохраняем старые значения для логирования изменений
        old_quantity = material.quantity_kg
        old_batch_number = material.batch_number
        new_quantity = form.quantity_kg.data
        new_batch_number = form.batch_number.data
        
        material.type_id = form.type_id.data
        material.batch_number = new_batch_number
        material.quantity_kg = new_quantity
        material.date_received = form.date_received.data or material.date_received
        material.expiration_date = form.expiration_date.data or material.expiration_date
        
        try:
            db.session.commit()
            
            # Синхронизируем номера партий в связанных записях
            if old_batch_number != new_batch_number:
                updated_count = sync_material_batch_numbers(material.id)
                if updated_count > 0:
                    flash(f'Номер партии синхронизирован в {updated_count} замесах', 'info')
            
            # Показываем информацию об изменении количества
            if old_quantity != new_quantity:
                change = new_quantity - old_quantity
                if change > 0:
                    flash(f'Партия сырья обновлена! Количество увеличено на {change:.2f} кг (с {old_quantity:.2f} до {new_quantity:.2f} кг)', 'success')
                else:
                    flash(f'Партия сырья обновлена! Количество уменьшено на {abs(change):.2f} кг (с {old_quantity:.2f} до {new_quantity:.2f} кг)', 'success')
            else:
                flash('Партия сырья обновлена!', 'success')
                
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при обновлении: {e}', 'error')
            
        return redirect(url_for('raw_materials'))
    return render_template('edit_raw_material.html', form=form, material=material)

@app.route('/raw_materials/delete/<int:id>', methods=['POST'])
@admin_required
def delete_raw_material(id):
    material = RawMaterial.query.get_or_404(id)
    db.session.delete(material)
    db.session.commit()
    flash('Партия сырья удалена!', 'success')
    return redirect(url_for('raw_materials'))

@app.route('/raw_materials/edit_used_up/<int:material_id>', methods=['POST'])
@admin_required
def edit_used_up_material(material_id):
    """Редактирует количество выработанного сырья с автоматическим перемещением между категориями"""
    material = RawMaterial.query.get_or_404(material_id)
    
    try:
        new_quantity = float(request.form.get('new_quantity', 0))
        
        if new_quantity < 0:
            flash('Количество не может быть отрицательным', 'error')
            return redirect(url_for('raw_materials'))
        
        old_quantity = material.quantity_kg
        
        # Обновляем количество
        material.quantity_kg = new_quantity
        
        # Логирование изменений (без поля notes, так как его нет в модели)
        timestamp = datetime.now().strftime('%d.%m.%Y %H:%M')
        print(f"[{timestamp}] Сырьё ID {material.id} ({material.batch_number}): количество изменено с {old_quantity:.2f} кг на {new_quantity:.2f} кг")
        
        db.session.commit()
        
        # Определяем тип сообщения в зависимости от изменения
        if old_quantity == 0 and new_quantity > 0:
            flash(f'Сырьё "{material.batch_number}" перемещено в доступное! Количество: {new_quantity:.2f} кг', 'success')
        elif old_quantity > 0 and new_quantity == 0:
            flash(f'Сырьё "{material.batch_number}" перемещено в выработанное!', 'info')
        else:
            flash(f'Количество сырья "{material.batch_number}" изменено с {old_quantity:.2f} кг на {new_quantity:.2f} кг', 'success')
        
    except ValueError:
        flash('Некорректное значение количества', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при обновлении: {e}', 'error')
    
    return redirect(url_for('raw_materials'))

# --- Продукты ---
@app.route('/products', methods=['GET', 'POST'])
@login_required
def products():
    form = ProductForm()
    if form.validate_on_submit():
        # Проверяем права на создание
        if not current_user.is_admin():
            flash('Доступ запрещен. Требуются права администратора для создания.', 'error')
            return redirect(url_for('products'))
        
        product = Product(name=form.name.data)
        db.session.add(product)
        db.session.commit()
        flash('Продукт добавлен!', 'success')
        return redirect(url_for('products'))
    products = Product.query.all()
    return render_template('products.html', form=form, products=products)

@app.route('/products/delete/<int:id>', methods=['POST'])
@admin_required
def delete_product(id):
    product = Product.query.get_or_404(id)
    
    # Проверяем, есть ли связанные рецептуры
    if product.recipe_templates:
        flash('Невозможно удалить продукт, так как для него существуют рецептуры.', 'error')
        return redirect(url_for('products'))
    
    # Проверяем, есть ли связанные планы производства
    if product.production_plans:
        flash('Невозможно удалить продукт, так как для него существуют планы производства.', 'error')
        return redirect(url_for('products'))
    
    db.session.delete(product)
    db.session.commit()
    flash('Продукт удален!', 'success')
    return redirect(url_for('products'))

# --- Рецептуры ---
@app.route('/recipes', methods=['GET', 'POST'])
@login_required
def recipes():
    form = RecipeForm()
    form.product_id.choices = [(p.id, p.name) for p in Product.query.all()]
    if form.validate_on_submit():
        # Проверяем права на создание
        if not current_user.is_admin():
            flash('Доступ запрещен. Требуются права администратора для создания.', 'error')
            return redirect(url_for('recipes'))
        
        recipe = Recipe(
            name=form.name.data, 
            product_id=form.product_id.data,
            status='draft'
        )
        db.session.add(recipe)
        db.session.commit()
        flash('Рецептура создана! Теперь добавьте ингредиенты.', 'success')
        return redirect(url_for('recipe_ingredients', recipe_id=recipe.id))
    recipes = Recipe.query.filter_by(status='saved').all()  # Показываем только сохраненные
    drafts = Recipe.query.filter_by(status='draft').all()  # Отдельно показываем черновики
    return render_template('recipes.html', form=form, recipes=recipes, drafts=drafts)

@app.route('/recipes/delete/<int:id>', methods=['POST'])
@admin_required
def delete_recipe(id):
    recipe = Recipe.query.get_or_404(id)
    
    # Проверяем, используется ли рецептура в планах производства
    plans_using_recipe = ProductionPlan.query.filter_by(template_id=recipe.id).all()
    if plans_using_recipe:
        plan_numbers = [str(plan.id) for plan in plans_using_recipe]
        flash(f'Невозможно удалить рецептуру, так как она используется в планах производства: {", ".join(plan_numbers)}', 'error')
        return redirect(url_for('recipes'))
    
    db.session.delete(recipe)
    db.session.commit()
    flash('Рецептура удалена!', 'success')
    return redirect(url_for('recipes'))

# --- Ингредиенты рецептуры ---
@app.route('/recipes/<int:recipe_id>/ingredients', methods=['GET', 'POST'])
@login_required
def recipe_ingredients(recipe_id):
    recipe = Recipe.query.get_or_404(recipe_id)
    
    # Для сохраненных рецептур запрещаем только модификации
    if recipe.status == 'saved' and request.method == 'POST':
        flash('Эта рецептура уже сохранена и не может быть изменена.', 'warning')
        return redirect(url_for('recipe_ingredients', recipe_id=recipe_id))
        
    form = RecipeIngredientForm()
    form.recipe = recipe  # для валидации
    
    # Получаем список типов сырья
    raw_material_types = RawMaterialType.query.all()
    form.material_type_id.choices = [(rmt.id, rmt.name) for rmt in raw_material_types]
    
    if request.method == 'POST':
        # Проверяем права на модификацию
        if not current_user.is_admin():
            flash('Доступ запрещен. Требуются права администратора для изменения рецептуры.', 'error')
            return redirect(url_for('recipe_ingredients', recipe_id=recipe_id))
        
        if 'save_recipe' in request.form:  # Нажата кнопка "Сохранить рецептуру"
            total_percent = sum(item.percentage for item in recipe.recipe_items)
            if not recipe.recipe_items:
                flash('Добавьте хотя бы один ингредиент перед сохранением рецептуры.', 'error')
            elif abs(total_percent - 100) > 0.01:  # Учитываем возможную погрешность float
                flash(f'Сумма процентов должна быть равна 100%. Текущая сумма: {total_percent:.3f}%', 'error')
            else:
                recipe.status = 'saved'
                db.session.commit()
                flash('Рецептура успешно сохранена!', 'success')
                return redirect(url_for('recipes'))
        elif form.validate():  # Нажата кнопка "Добавить ингредиент"
            # Проверяем, не добавлен ли уже этот тип сырья
            existing = RecipeIngredient.query.filter_by(
                template_id=recipe.id,
                material_type_id=form.material_type_id.data
            ).first()
            
            if existing:
                flash('Этот вид сырья уже добавлен в рецептуру.', 'error')
            else:
                ingredient = RecipeIngredient(
                    template_id=recipe.id,
                    material_type_id=form.material_type_id.data,
                    percentage=form.percentage.data
                )
                db.session.add(ingredient)
                db.session.commit()
                flash('Ингредиент добавлен!', 'success')
                
                # Показываем текущую сумму процентов
                total_percent = sum(item.percentage for item in recipe.recipe_items)
                if total_percent > 100:
                    flash(f'Предупреждение: сумма процентов превышает 100% ({total_percent:.3f}%)', 'warning')
                elif total_percent < 100:
                    flash(f'Осталось добавить {100 - total_percent:.3f}%', 'info')
            
        return redirect(url_for('recipe_ingredients', recipe_id=recipe.id))
        
    ingredients = RecipeIngredient.query.filter_by(template_id=recipe.id).all()
    total_percent = sum(item.percentage for item in ingredients)
    return render_template(
        'recipe_ingredients.html',
        form=form if recipe.status != 'saved' else None,  # Не показываем форму для сохраненных рецептур
        recipe=recipe,
        ingredients=ingredients,
        total_percent=total_percent
    )

@app.route('/recipes/<int:recipe_id>/ingredients/<int:ingredient_id>/delete', methods=['POST'])
@admin_required
def delete_recipe_ingredient(recipe_id, ingredient_id):
    recipe = Recipe.query.get_or_404(recipe_id)
    if recipe.status == 'saved':
        flash('Нельзя изменять сохраненную рецептуру.', 'error')
        return redirect(url_for('recipes'))
        
    ingredient = RecipeIngredient.query.get_or_404(ingredient_id)
    if ingredient.template_id != recipe_id:
        flash('Ошибка: ингредиент не принадлежит указанной рецептуре.', 'error')
        return redirect(url_for('recipes'))
        
    db.session.delete(ingredient)
    db.session.commit()
    flash('Ингредиент удален из рецептуры!', 'success')
    return redirect(url_for('recipe_ingredients', recipe_id=recipe_id))

@app.route('/production_plans')
@login_required
def production_plans():
    # Получаем параметры фильтрации
    product_id = request.args.get('product_id', type=int)
    status = request.args.get('status')
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')

    query = ProductionPlan.query

    # Применяем фильтры
    if product_id:
        query = query.filter(ProductionPlan.product_id == product_id)
    if status:
        query = query.filter(ProductionPlan.status == status)
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
            query = query.filter(ProductionPlan.created_at >= date_from)
        except ValueError:
            flash('Неверный формат даты начала периода.', 'error')
    
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
            # Включаем весь день до 23:59:59
            query = query.filter(ProductionPlan.created_at <= date_to + timedelta(days=1, seconds=-1))
        except ValueError:
            flash('Неверный формат даты конца периода.', 'error')

    # Сортировка
    plans = query.order_by(ProductionPlan.created_at.desc()).all()

    # Данные для форм фильтров
    products = Product.query.order_by(Product.name).all()

    # Цвета для статусов
    status_colors = {
        PlanStatus.DRAFT: 'secondary',
        PlanStatus.APPROVED: 'info',
        PlanStatus.IN_PROGRESS: 'primary',
        PlanStatus.COMPLETED: 'success',
        PlanStatus.CANCELLED: 'danger'
    }

    return render_template(
        'production_plans.html',
        plans=plans,
        products=products,
        status_colors=status_colors,
        PlanStatus=PlanStatus
    )

@app.route('/production_plans/<int:plan_id>/update_status', methods=['POST'])
@operator_required
def update_plan_status(plan_id):
    plan = ProductionPlan.query.get_or_404(plan_id)
    form = ProductionStatusForm(plan=plan)
    
    if form.validate_on_submit():
        old_status = plan.status
        new_status = form.status.data

        # Проверка перед установкой статуса "завершен"
        if new_status == PlanStatus.COMPLETED:
            # Проверяем, что указана дата производства
            # if not form.production_date.data:  # Временно отключено
            #     flash('Для завершения плана необходимо указать дату производства', 'error')
            #     return redirect(url_for('production_plan_detail', plan_id=plan.id))
            
            total_produced = sum(batch.weight for batch in plan.batches)
            # Сравниваем с небольшой погрешностью для чисел с плавающей точкой
            if total_produced < plan.quantity * 0.999:
                flash(f'Невозможно завершить план. Произведено {total_produced:.2f} кг из {plan.quantity:.2f} кг.', 'error')
                return redirect(url_for('production_plan_detail', plan_id=plan.id))
            # Строгая проверка внесения всех ингредиентов
            for batch in plan.batches:
                for recipe_item in plan.template.recipe_items:
                    need_qty = batch.weight * float(recipe_item.percentage) / 100
                    added_qty = sum(
                        bi.quantity for bi in batch.materials
                        if bi.material_batch.material.type_id == recipe_item.material_type_id
                    )
                    if abs(added_qty - need_qty) > 0.01:  # допускаем небольшую погрешность
                        flash(
                            f'Невозможно завершить план. В замесе №{batch.batch_number} для ингредиента "{recipe_item.material_type.name}" внесено {added_qty:.2f} кг из {need_qty:.2f} кг.',
                            'error'
                        )
                        return redirect(url_for('production_plan_detail', plan_id=plan.id))
        
        # Списание сырья при завершении плана
        if old_status != PlanStatus.COMPLETED and new_status == PlanStatus.COMPLETED:
            try:
                for batch in plan.batches:
                    for batch_ingredient in batch.materials:
                        raw_material = batch_ingredient.material_batch.material
                        if raw_material:
                            raw_material.quantity_kg = max(0, raw_material.quantity_kg - batch_ingredient.quantity)
                db.session.commit()
                flash('Сырьё списано при завершении плана', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка при списании сырья: {str(e)}', 'error')
                return redirect(url_for('production_plan_detail', plan_id=plan.id))
        
        # Обновляем дату производства только если план не завершён
        # if new_status != PlanStatus.COMPLETED and form.production_date.data:  # Временно отключено
        #     old_date = plan.production_date
        #     plan.production_date = form.production_date.data
        #     
        #     # Логируем изменение даты производства
        #     if old_date != plan.production_date:
        #         timestamp = datetime.now().strftime('%d.%m.%Y %H:%M')
        #         if old_date:
        #             date_note = f"[{timestamp}] Дата производства изменена: {old_date.strftime('%d.%m.%Y')} → {plan.production_date.strftime('%d.%m.%Y')}"
        #         else:
        #             date_note = f"[{timestamp}] Дата производства установлена: {plan.production_date.strftime('%d.%m.%Y')}"
        #         
        #         if plan.notes:
        #             plan.notes = date_note + "\n\n" + plan.notes
        #         else:
        #             plan.notes = date_note
        
        # Обработка старого формата статуса (без ё)
        old_status_normalized = old_status.replace("утвержден", "утверждён")
        
        # Получаем отображаемые значения статусов, с защитой от отсутствующих ключей
        old_status_display = old_status_normalized
        new_status_display = new_status
        
        timestamp = datetime.now().strftime('%d.%m.%Y %H:%M')
        status_note = f"[{timestamp}] Статус изменен: {old_status_display} → {new_status_display}"
        
        # Добавляем примечание к существующим
        if plan.notes:
            plan.notes = status_note + "\n" + plan.notes
        else:
            plan.notes = status_note
            
        if form.notes.data:
            plan.notes = form.notes.data + "\n" + plan.notes
        
        plan.status = new_status
        db.session.commit()
        
        flash('Статус успешно обновлен!', 'success')
        return redirect(url_for('production_plan_detail', plan_id=plan.id))
    
    return redirect(url_for('production_plan_detail', plan_id=plan.id))

def get_actual_available_quantity(raw_material):
    """
    Вычисляет реальный остаток сырья с учётом уже использованных количеств в замесах.
    """
    # Получаем сумму всех использованных количеств этого сырья в замесах
    # НО только для незавершённых планов (где сырьё ещё не списано)
    used_quantity = db.session.query(db.func.sum(BatchMaterial.quantity)).join(
        MaterialBatch, BatchMaterial.material_batch_id == MaterialBatch.id
    ).join(
        ProductionBatch, BatchMaterial.batch_id == ProductionBatch.id
    ).join(
        ProductionPlan, ProductionBatch.plan_id == ProductionPlan.id
    ).filter(
        MaterialBatch.material_id == raw_material.id,
        ProductionPlan.status != 'завершен'  # Не учитываем завершённые планы
    ).scalar() or 0
    
    # Реальный остаток = общее количество - использованное
    actual_available = raw_material.quantity_kg - used_quantity
    return max(0, actual_available)  # Не может быть отрицательным

def get_available_materials_for_batch(needed_qty, material_type_id):
    """
    Получает список партий сырья с учётом остатков и срока годности.
    Возвращает список партий с количеством для использования и оставшуюся потребность.
    """
    materials = RawMaterial.query.filter_by(type_id=material_type_id).filter(
        RawMaterial.quantity_kg > 0
    ).order_by(
        RawMaterial.expiration_date.asc().nullslast()
    ).all()
    
    result = []
    remaining_qty = needed_qty
    
    for material in materials:
        if remaining_qty <= 0:
            break
            
        # Используем реальный остаток с учётом уже использованных количеств
        available_qty = get_actual_available_quantity(material)
        qty_to_use = min(available_qty, remaining_qty)
        
        if qty_to_use > 0:
            result.append({
                'material': material,
                'quantity': qty_to_use
            })
            remaining_qty -= qty_to_use
    
    return result, remaining_qty

@app.route('/production_plans/<int:plan_id>/add_batch', methods=['POST'])
@operator_required
def add_batch(plan_id):
    plan = ProductionPlan.query.get_or_404(plan_id)
    form = ProductionBatchForm()
    
    if form.validate_on_submit():
        # Проверяем максимальный вес замеса
        if form.quantity.data > 1000:
            flash('Вес замеса не может превышать 1000 кг', 'error')
            return redirect(url_for('production_plan_detail', plan_id=plan_id))
        
        # Проверяем уникальность номера замеса в рамках текущего плана
        existing_batch = ProductionBatch.query.filter_by(
            plan_id=plan.id,
            batch_number=form.batch_number.data
        ).first()
        if existing_batch:
            flash('Замес с таким номером уже существует в текущем плане', 'error')
            return redirect(url_for('production_plan_detail', plan_id=plan_id))
        
        # Проверяем, не превышает ли общее количество план
        total_quantity = sum([batch.weight for batch in plan.batches]) + form.quantity.data
        if total_quantity > plan.quantity:
            flash(f'Общее количество замесов ({total_quantity} кг) превышает план ({plan.quantity} кг)', 'error')
            return redirect(url_for('production_plan_detail', plan_id=plan_id))
        
        batch = ProductionBatch(
            plan=plan,
            batch_number=form.batch_number.data,
            weight=form.quantity.data
        )
        db.session.add(batch)
        db.session.flush()  # Получаем ID замеса
        
        # Автоматически добавляем все ингредиенты из рецептуры
        added_ingredients = []
        missing_ingredients = []
        
        for recipe_item in plan.template.recipe_items:
            needed_qty = batch.weight * float(recipe_item.percentage) / 100
            
            # Получаем список партий с нужным количеством
            materials_to_use, shortage = get_available_materials_for_batch(
                needed_qty, recipe_item.material_type_id
            )
            
            if shortage > 0:
                missing_ingredients.append(f"{recipe_item.material_type.name} (нужно {needed_qty:.2f} кг, недостаточно {shortage:.2f} кг)")
                continue
            
            # Создаём записи для каждой использованной партии
            for material_info in materials_to_use:
                material = material_info['material']
                qty = material_info['quantity']
                
                # Создаём MaterialBatch и BatchMaterial
                material_batch = MaterialBatch(
                    material=material,
                    batch_number=material.batch_number,
                    quantity=qty,
                    remaining_quantity=qty
                )
                db.session.add(material_batch)
                db.session.flush()
                
                ingredient = BatchMaterial(
                    batch=batch,
                    material_batch=material_batch,
                    quantity=qty
                )
                db.session.add(ingredient)
            
            # Добавляем информацию об ингредиенте
            total_used = sum(m['quantity'] for m in materials_to_use)
            if len(materials_to_use) == 1:
                added_ingredients.append(f"{recipe_item.material_type.name} ({total_used:.2f} кг из партии {materials_to_use[0]['material'].batch_number})")
            else:
                batches_info = [f"{m['material'].batch_number}({m['quantity']:.2f} кг)" for m in materials_to_use]
                added_ingredients.append(f"{recipe_item.material_type.name} ({total_used:.2f} кг из партий: {', '.join(batches_info)})")
        
        # Добавляем запись в примечания
        timestamp = datetime.now().strftime('%d.%m.%Y %H:%M')
        batch_note = f"[{timestamp}] Добавлен замес №{batch.batch_number} ({batch.weight} кг)"
        
        if added_ingredients:
            batch_note += f"\nАвтоматически добавлены ингредиенты"
        
        if missing_ingredients:
            batch_note += f"\nНедостаточно сырья: {', '.join(missing_ingredients)}"
        
        if plan.notes:
            plan.notes = batch_note + "\n\n" + plan.notes
        else:
            plan.notes = batch_note
            
        db.session.commit()
        
        if missing_ingredients:
            flash(f'Замес добавлен. Недостаточно сырья: {", ".join(missing_ingredients)}', 'warning')
        else:
            flash('Замес добавлен с автоматическим заполнением ингредиентов', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Ошибка в поле {getattr(form, field).label.text}: {error}', 'error')
    
    return redirect(url_for('production_plan_detail', plan_id=plan_id))

@app.route('/production_plans/<int:plan_id>/add_batch_ingredient', methods=['GET', 'POST'])
@operator_required
def add_batch_ingredient(plan_id):
    form = BatchIngredientForm()
    batch_id = request.args.get('batch_id') or request.form.get('batch_id')
    ingredient_type_id = request.args.get('ingredient_type_id') or request.form.get('ingredient_type_id')
    if not batch_id or not ingredient_type_id:
        flash('Не выбран замес или ингредиент для добавления', 'error')
        return redirect(url_for('production_plan_detail', plan_id=plan_id))
    batch = ProductionBatch.query.get_or_404(batch_id)
    ingredient_info = next((info for info in batch.plan.template.recipe_items if str(info.material_type_id) == str(ingredient_type_id)), None)
    if not ingredient_info:
        flash('Выбран некорректный ингредиент', 'error')
        return redirect(url_for('production_plan_detail', plan_id=plan_id))
    available_materials = RawMaterial.query.filter_by(type_id=ingredient_info.material_type_id).all()
    form.raw_material_id.choices = [(m.id, f"{m.batch_number} ({m.quantity_kg} кг)") for m in available_materials]
    if form.validate_on_submit():
        raw_material = RawMaterial.query.get(form.raw_material_id.data)
        if not raw_material or raw_material.type_id != ingredient_info.material_type_id:
            flash('Выбрано неверное сырье', 'error')
            return redirect(url_for('production_plan_detail', plan_id=plan_id))
        need_qty = (batch.weight * float(ingredient_info.percentage) / 100) - sum([
            bi.quantity for bi in batch.materials
            if bi.material_batch.material.type_id == ingredient_info.material_type_id
        ])
        if form.quantity.data > need_qty:
            flash(f'Нельзя добавить больше чем нужно ({need_qty:.2f} кг)', 'error')
            return redirect(url_for('production_plan_detail', plan_id=plan_id))
        if form.quantity.data > raw_material.quantity_kg:
            flash(f'Недостаточно сырья (доступно {raw_material.quantity_kg:.2f} кг)', 'error')
            return redirect(url_for('production_plan_detail', plan_id=plan_id))
        material_batch = MaterialBatch(
            material=raw_material,
            batch_number=raw_material.batch_number,
            quantity=form.quantity.data,
            remaining_quantity=form.quantity.data
        )
        db.session.add(material_batch)
        db.session.flush()
        ingredient = BatchMaterial(
            batch=batch,
            material_batch=material_batch,
            quantity=form.quantity.data
        )
        db.session.add(ingredient)
        db.session.commit()
        flash('Ингредиент добавлен!', 'success')
        return redirect(url_for('production_plan_detail', plan_id=plan_id))
    return render_template('add_batch_ingredient.html', form=form, batch=batch)

@app.route('/production_plans/<int:plan_id>')
@login_required
def production_plan_detail(plan_id):
    plan = ProductionPlan.query.get_or_404(plan_id)
    
    # Проверяем существование рецептуры
    if not plan.template:
        flash('Рецептура для данного плана была удалена. Измените статус плана на "черновик" или "отменён", чтобы удалить его.', 'warning')
    
    # Вычисляем прогресс производства
    total_produced = sum([batch.weight for batch in plan.batches])
    progress_percent = (total_produced / plan.quantity * 100) if plan.quantity > 0 else 0
    
    # Проверяем наличие сырья только если есть рецептура
    raw_materials_availability = {}
    can_start = True
    start_message = ""
    
    if plan.template:
        for ingredient in plan.template.recipe_items:
            type_id = ingredient.material_type_id
            needed_qty = (plan.quantity * float(ingredient.percentage) / 100)
            available_qty = sum([rm.quantity_kg for rm in RawMaterial.query.filter_by(type_id=type_id).all()])
            
            raw_materials_availability[type_id] = {
                'type': ingredient.material_type,
                'quantity_needed': needed_qty,
                'quantity_available': available_qty,
                'is_available': available_qty >= needed_qty,
                'shortage': max(0, needed_qty - available_qty)
            }
            
            if available_qty < needed_qty:
                can_start = False
                start_message = "Недостаточно сырья для начала производства"

    # Цвета для статусов
    status_colors = {
        PlanStatus.DRAFT: 'secondary',
        PlanStatus.APPROVED: 'info',
        PlanStatus.IN_PROGRESS: 'primary',
        PlanStatus.COMPLETED: 'success',
        PlanStatus.CANCELLED: 'danger'
    }

    # Информация о замесах
    batches_info = []
    if plan.template:
        for batch in plan.batches:
            ingredients_info = []
            for recipe_ingredient in plan.template.recipe_items:
                type_id = recipe_ingredient.material_type_id
                need_qty = batch.weight * float(recipe_ingredient.percentage) / 100
                batch_ingredients_query = BatchMaterial.query.filter_by(batch_id=batch.id)
                batch_ingredients = [
                    bi for bi in batch_ingredients_query.all()
                    if bi.material_batch.material.type_id == type_id
                ]
                added_qty = sum(bi.quantity for bi in batch_ingredients)

                # Группируем партии по batch_number
                grouped_batches = {}
                for bi in batch_ingredients:
                    key = bi.material_batch.batch_number
                    if key not in grouped_batches:
                        grouped_batches[key] = {
                            'batch_number': bi.material_batch.batch_number,
                            'material_name': bi.material_batch.material.name,
                            'quantity': 0.0
                        }
                    grouped_batches[key]['quantity'] += bi.quantity

                ingredients_info.append({
                    'type': recipe_ingredient.material_type,
                    'need_qty': need_qty,
                    'added_qty': added_qty,
                    'to_add_qty': max(0, need_qty - added_qty),
                    'added_materials': list(grouped_batches.values()),
                    'original_batch_materials': batch_ingredients
                })
            
            batches_info.append({
                'batch': batch,
                'ingredients': ingredients_info
            })

    # Создаем формы
    status_form = ProductionStatusForm(plan=plan)
    status_form.status.data = plan.status  # Устанавливаем текущий статус
    batch_form = ProductionBatchForm()
    batch_ingredient_form = BatchIngredientForm()

    return render_template(
        'production_plan_detail.html',
        plan=plan,
        raw_materials_availability=raw_materials_availability,
        can_start=can_start,
        start_message=start_message,
        progress_percent=progress_percent,
        total_produced=total_produced,
        status_colors=status_colors,
        batches_info=batches_info,
        status_form=status_form,
        batch_form=batch_form,
        batch_ingredient_form=batch_ingredient_form,
        PlanStatus=PlanStatus)

@app.route('/batches/delete/<int:batch_id>', methods=['POST'])
@login_required
def delete_batch(batch_id):
    batch = ProductionBatch.query.get_or_404(batch_id)
    plan = batch.plan
    if plan.status == 'completed':
        flash('Нельзя удалять замес после завершения плана!', 'danger')
        return redirect(url_for('production_plan_detail', plan_id=plan.id))
    db.session.delete(batch)
    db.session.commit()
    flash('Замес удалён!', 'success')
    return redirect(url_for('production_plan_detail', plan_id=plan.id))

@app.route('/batch_ingredients/delete/<int:ingredient_id>', methods=['POST'])
@login_required
def delete_batch_ingredient(ingredient_id):
    # Находим ингредиент, который нужно удалить
    ingredient = BatchMaterial.query.get_or_404(ingredient_id)
    plan_id = ingredient.batch.plan_id

    # Проверяем статус плана, чтобы запретить изменения в завершенных планах
    if ingredient.batch.plan.status == PlanStatus.COMPLETED:
        flash('Нельзя удалять ингредиенты из завершенного плана!', 'danger')
        return redirect(url_for('production_plan_detail', plan_id=plan_id))

    try:
        # Удаляем запись об использовании ингредиента
        db.session.delete(ingredient)
        db.session.commit()
        flash('Ингредиент удален из замеса!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при удалении ингредиента: {str(e)}', 'error')
        
    return redirect(url_for('production_plan_detail', plan_id=plan_id))

@app.route('/reports')
@login_required
def reports():
    # Остатки сырья
    raw_materials = RawMaterial.query.order_by(RawMaterial.type_id, RawMaterial.batch_number).all()
    # История производства
    plans = ProductionPlan.query.order_by(ProductionPlan.created_at.desc()).all()
    return render_template('reports.html', raw_materials=raw_materials, plans=plans)

@app.route('/recipes/by_product/<int:product_id>')
@login_required
def recipes_by_product(product_id):
    recipes = Recipe.query.filter_by(
        product_id=product_id,
        status='saved'
    ).all()
    return jsonify([{
        'id': recipe.id,
        'name': recipe.name
    } for recipe in recipes])

@app.route('/raw_materials/by_type/<int:type_id>')
@login_required
def raw_materials_by_type(type_id):
    materials = RawMaterial.query.filter_by(type_id=type_id).all()
    return jsonify([{
        'id': m.id,
        'type_name': m.material_type.name,
        'batch_number': m.batch_number,
        'quantity_kg': m.quantity_kg
    } for m in materials])

@app.route('/raw_materials/available/<int:type_id>')
@login_required
def get_available_raw_materials(type_id):
    materials = RawMaterial.query.filter_by(type_id=type_id).all()
    return jsonify([{
        'id': m.id,
        'name': f"{m.batch_number} ({m.quantity_kg:.2f} кг)"
    } for m in materials if m.quantity_kg > 0])

@app.route('/production_plans/create', methods=['GET', 'POST'])
@admin_required
def create_production_plan():
    form = ProductionPlanForm()
    
    # Заполняем список продуктов
    products = Product.query.order_by(Product.name).all()
    form.product_id.choices = [(p.id, p.name) for p in products]
    
    # Заполняем список рецептур для выбранного продукта
    if request.method == 'POST':
        selected_product_id = int(request.form.get('product_id', 0))
    else:
        selected_product_id = request.args.get('product_id', type=int)
        if not selected_product_id and products:
            selected_product_id = products[0].id
    
    if selected_product_id:
        recipes = Recipe.query.filter_by(
            product_id=selected_product_id,
            status='saved'
        ).order_by(Recipe.name).all()
        form.template_id.choices = [(r.id, r.name) for r in recipes]
    else:
        form.template_id.choices = []
    
    if form.validate_on_submit():
        plan = ProductionPlan(
            product_id=form.product_id.data,
            template_id=form.template_id.data,
            batch_number=form.batch_number.data,
            quantity=form.quantity.data,
            notes=form.notes.data,
            status=PlanStatus.DRAFT
        )
        
        # Если нажата кнопка "Утвердить"
        if form.approve.data:
            # Проверка наличия сырья уже выполнена в validate_quantity
            plan.status = PlanStatus.APPROVED
            flash('План производства создан и утверждён!', 'success')
        else:
            flash('План производства создан как черновик', 'success')
        
        db.session.add(plan)
        try:
            db.session.commit()
            return redirect(url_for('production_plan_detail', plan_id=plan.id))
        except Exception as e:
            db.session.rollback()
            if 'UNIQUE constraint' in str(e):
                flash('План производства с таким номером партии уже существует', 'error')
            else:
                flash('Произошла ошибка при сохранении плана', 'error')
            return redirect(url_for('create_production_plan'))
    
    return render_template('create_production_plan.html',
                         form=form,
                         selected_product_id=selected_product_id)

@app.route('/production_plans/<int:plan_id>/delete', methods=['POST'])
@admin_required
def delete_production_plan(plan_id):
    plan = ProductionPlan.query.get_or_404(plan_id)
    
    # Проверяем, можно ли удалить план
    if plan.status not in [PlanStatus.DRAFT, PlanStatus.CANCELLED]:
        flash('Можно удалить только черновик плана производства или отмененный план', 'error')
        return redirect(url_for('production_plans'))
    
    # Получаем номер партии для сообщения
    batch_number = plan.batch_number or f'#{plan.id}'
    
    try:
        # Удаляем все замесы и их ингредиенты
        for batch in plan.batches:
            db.session.delete(batch)
        
        # Удаляем сам план
        db.session.delete(plan)
        db.session.commit()
        
        flash(f'План производства {batch_number} удален', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при удалении плана производства: {str(e)}', 'error')
    
    return redirect(url_for('production_plans'))

@app.route('/reports/raw_material_usage', methods=['GET', 'POST'])
@login_required
def raw_material_usage_report():
    form = RawMaterialUsageReportForm()
    usage_data = None

    if form.validate_on_submit():
        date_from = form.date_from.data
        date_to = form.date_to.data

        query = (
            db.session.query(
                RawMaterialType.name.label('type_name'),
                MaterialBatch.batch_number,
                db.func.sum(BatchMaterial.quantity).label('used_qty')
            )
            .join(MaterialBatch, MaterialBatch.id == BatchMaterial.material_batch_id)
            .join(RawMaterial, RawMaterial.id == MaterialBatch.material_id)
            .join(RawMaterialType, RawMaterialType.id == RawMaterial.type_id)
            .join(ProductionBatch, ProductionBatch.id == BatchMaterial.batch_id)
            .join(ProductionPlan, ProductionPlan.id == ProductionBatch.plan_id)
            .filter(BatchMaterial.quantity > 0)
        )
        if date_from:
            query = query.filter(ProductionPlan.created_at >= date_from)
        if date_to:
            query = query.filter(ProductionPlan.created_at <= date_to)
        query = query.group_by(RawMaterialType.name, MaterialBatch.batch_number)
        usage_data = query.all()
    else:
        usage_data = None

    return render_template(
        'raw_material_usage_report.html',
        form=form,
        usage_data=usage_data
    )

@app.route('/reports/production_statistics', methods=['GET', 'POST'])
@login_required
def production_statistics():
    form = ProductionStatisticsForm()
    statistics_data = None
    totals = {}

    if form.validate_on_submit():
        date_from = form.date_from.data
        date_to = form.date_to.data

        query = (
            ProductionPlan.query
            .options(
                joinedload(ProductionPlan.batches)
                .joinedload(ProductionBatch.materials)
                .joinedload(BatchMaterial.material_batch)
                .joinedload(MaterialBatch.material)
                .joinedload(RawMaterial.type)
            )
            .join(Product, Product.id == ProductionPlan.product_id)
            .join(Recipe, Recipe.id == ProductionPlan.template_id)
            .filter(ProductionPlan.status == PlanStatus.COMPLETED)
        )

        if date_from:
            query = query.filter(ProductionPlan.created_at >= date_from)
        if date_to:
            date_to_end_of_day = date_to + timedelta(days=1, seconds=-1)
            query = query.filter(ProductionPlan.created_at <= date_to_end_of_day)

        # Сортировка по дате и продукту
        statistics_data = query.order_by(Product.name, ProductionPlan.created_at).all()
        
        # Подготовим итоговые данные по продуктам
        if statistics_data:
            for plan in statistics_data:
                key = (plan.product.name, plan.template.name)
                if key not in totals:
                    totals[key] = {
                        'product_name': plan.product.name,
                        'recipe_name': plan.template.name,
                        'total_quantity': 0,
                        'plans_count': 0
                    }
                totals[key]['total_quantity'] += plan.quantity
                totals[key]['plans_count'] += 1

    return render_template(
        'production_statistics.html',
        form=form,
        statistics_data=statistics_data,
        totals=totals if statistics_data else None
    )

@app.route('/reports/raw_material_forecast', methods=['GET'])
@login_required
def raw_material_forecast():
    # Получаем все утвержденные планы производства
    approved_plans = ProductionPlan.query.filter_by(status=PlanStatus.APPROVED).order_by(ProductionPlan.created_at).all()
    
    # Получаем все типы сырья
    raw_material_types = RawMaterialType.query.all()
    
    # Текущие остатки по типам сырья
    current_stock = {}
    for material in RawMaterial.query.all():
        if material.type_id not in current_stock:
            current_stock[material.type_id] = 0
        current_stock[material.type_id] += material.quantity_kg
    
    # Прогноз потребности в сырье по дням
    forecast = {}
    for plan in approved_plans:
        date_str = plan.created_at.strftime('%Y-%m-%d')
        if date_str not in forecast:
            forecast[date_str] = {t.id: 0 for t in raw_material_types}
        
        plan_quantity = plan.quantity
        recipe = plan.template
        
        for ingredient in recipe.recipe_items:
            needed_kg = (float(ingredient.percentage) / 100) * plan_quantity
            forecast[date_str][ingredient.material_type_id] += needed_kg
    
    # Накопительное использование по типам сырья
    cumulative_usage = {t.id: 0 for t in raw_material_types}
    for date_data in forecast.values():
        for type_id, usage in date_data.items():
            cumulative_usage[type_id] += usage
    
    # Сортируем даты для отображения
    sorted_dates = sorted(forecast.keys())
    
    return render_template(
        'raw_material_forecast.html',
        raw_material_types=raw_material_types,
        current_stock=current_stock,
        forecast=forecast,
        cumulative_usage=cumulative_usage,
        sorted_dates=sorted_dates
    )

@app.route('/reports/production_plans', methods=['GET'])
@login_required
def production_plans_report():
    """Отчёт по планам производства"""
    # Получаем параметры фильтрации
    product_id = request.args.get('product_id', type=int)
    status = request.args.get('status')
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')

    query = ProductionPlan.query

    # Применяем фильтры
    if product_id:
        query = query.filter(ProductionPlan.product_id == product_id)
    if status:
        query = query.filter(ProductionPlan.status == status)
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
            query = query.filter(ProductionPlan.created_at >= date_from)
        except ValueError:
            flash('Неверный формат даты начала периода.', 'error')
    
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
            # Включаем весь день до 23:59:59
            query = query.filter(ProductionPlan.created_at <= date_to + timedelta(days=1, seconds=-1))
        except ValueError:
            flash('Неверный формат даты конца периода.', 'error')

    # Сортировка
    plans = query.order_by(ProductionPlan.created_at.desc()).all()

    # Данные для форм фильтров
    products = Product.query.order_by(Product.name).all()

    # Цвета для статусов
    status_colors = {
        PlanStatus.DRAFT: 'secondary',
        PlanStatus.APPROVED: 'info',
        PlanStatus.IN_PROGRESS: 'primary',
        PlanStatus.COMPLETED: 'success',
        PlanStatus.CANCELLED: 'danger'
    }

    return render_template(
        'production_plans_report.html',
        plans=plans,
        products=products,
        status_colors=status_colors,
        PlanStatus=PlanStatus
    )

@app.route('/reports/raw_material_usage/export')
@login_required
def export_raw_material_usage():
    wb, output = create_excel_report()
    
    # Создаем лист для отчета
    ws = wb.create_sheet("Использование сырья")
    
    # Заголовки
    headers = ["Дата", "Вид сырья", "Партия", "Использовано (кг)", "План производства", "Продукт"]
    ws.append(headers)
    style_header_row(ws)
    
    # Получаем данные
    usage_data = []
    plans = ProductionPlan.query.order_by(ProductionPlan.created_at.desc()).all()
    
    for plan in plans:
        for batch in plan.batches:
            for ingredient in batch.materials:
                usage_data.append([
                    plan.created_at.strftime("%Y-%m-%d"),
                    ingredient.material_batch.material.type.name,
                    ingredient.material_batch.material.batch_number,
                    ingredient.quantity,
                    f"№{plan.batch_number}",
                    plan.product.name
                ])
    
    # Добавляем данные
    for row in usage_data:
        ws.append(row)
    
    # Форматирование
    adjust_column_width(ws)
    
    # Сохраняем и отправляем файл
    output = save_excel_report(wb, output)
    filename = f"raw_material_usage_{format_datetime(datetime.now())}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/reports/production_statistics/export')
@login_required
def export_production_statistics():
    wb, output = create_excel_report()
    
    # Создаем лист для отчета
    ws = wb.create_sheet("Детализация производства")
    
    # Заголовки
    headers = [
        "Дата плана", "Продукт", "Рецептура", "План №", "Статус плана", "Плановое кол-во (кг)",
        "Замес №", "Вес замеса (кг)",
        "Сырье", "Партия сырья", "Кол-во сырья (кг)"
    ]
    ws.append(headers)
    style_header_row(ws)
    
    # Получаем все данные одним запросом для эффективности
    plans = ProductionPlan.query.options(
        joinedload(ProductionPlan.product),
        joinedload(ProductionPlan.template),
        joinedload(ProductionPlan.batches).options(
            joinedload(ProductionBatch.materials).options(
                joinedload(BatchMaterial.material_batch).options(
                    joinedload(MaterialBatch.material).options(
                        joinedload(RawMaterial.type)
                    )
                )
            )
        )
    ).order_by(ProductionPlan.created_at.desc()).all()
    
    # Заполняем лист данными
    for plan in plans:
        plan_info = [
            plan.created_at.strftime("%Y-%m-%d"),
            plan.product.name,
            plan.template.name,
            plan.batch_number,
            plan.status,
            plan.quantity
        ]
        
        if not plan.batches:
            ws.append(plan_info) # Добавляем только информацию о плане, если нет замесов
        else:
            for batch in plan.batches:
                batch_info = plan_info + [
                    batch.batch_number,
                    batch.weight
                ]
                if not batch.materials:
                    ws.append(batch_info) # Добавляем инфо о плане и замесе, если нет ингредиентов
                else:
                    for material in batch.materials:
                        material_info = batch_info + [
                            material.material_batch.material.type.name,
                            material.material_batch.batch_number,
                            material.quantity
                        ]
                        ws.append(material_info)
                        
    adjust_column_width(ws)
    
    # Удаляем стандартный пустой лист, создаваемый библиотекой
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # Сохраняем и отправляем файл
    output = save_excel_report(wb, output)
    filename = f"production_statistics_{format_datetime(datetime.now())}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/reports/raw_material_forecast/export')
@login_required
def export_raw_material_forecast():
    wb, output = create_excel_report()
    
    # Текущие остатки
    ws_stock = wb.create_sheet("Текущие остатки")
    headers_stock = ["Вид сырья", "Остаток (кг)"]
    ws_stock.append(headers_stock)
    style_header_row(ws_stock)
    
    current_stock = {}
    for material in RawMaterial.query.all():
        if material.type_id not in current_stock:
            current_stock[material.type_id] = 0
        current_stock[material.type_id] += material.quantity_kg
    
    for type in RawMaterialType.query.all():
        ws_stock.append([
            type.name,
            current_stock.get(type.id, 0)
        ])
    
    adjust_column_width(ws_stock)
    
    # Прогноз по дням
    ws_forecast = wb.create_sheet("Прогноз по дням")
    raw_material_types = RawMaterialType.query.all()
    
    # Заголовки для прогноза
    headers_forecast = ["Дата"] + [t.name for t in raw_material_types]
    ws_forecast.append(headers_forecast)
    style_header_row(ws_forecast)
    
    # Получаем прогноз
    forecast = {}
    approved_plans = ProductionPlan.query.filter_by(status=PlanStatus.APPROVED).order_by(ProductionPlan.created_at).all()
    
    for plan in approved_plans:
        date_str = plan.created_at.strftime('%Y-%m-%d')
        if date_str not in forecast:
            forecast[date_str] = {t.id: 0 for t in raw_material_types}
        
        plan_quantity = plan.quantity
        recipe = plan.template
        
        for ingredient in recipe.recipe_items:
            needed_kg = (float(ingredient.percentage) / 100) * plan_quantity
            forecast[date_str][ingredient.material_type_id] += needed_kg
    
    # Добавляем данные прогноза
    for date in sorted(forecast.keys()):
        row = [date] + [forecast[date][t.id] for t in raw_material_types]
        ws_forecast.append(row)
    
    # Добавляем итоговую строку
    total_row = ["Общая потребность"]
    for type in raw_material_types:
        total = sum(forecast[date][type.id] for date in forecast)
        total_row.append(total)
    ws_forecast.append([])  # Пустая строка перед итогами
    ws_forecast.append(total_row)
    
    adjust_column_width(ws_forecast)
    
    # Анализ достаточности
    ws_analysis = wb.create_sheet("Анализ достаточности")
    headers_analysis = ["Вид сырья", "Текущий остаток (кг)", "Общая потребность (кг)", "Баланс (кг)", "Статус"]
    ws_analysis.append(headers_analysis)
    style_header_row(ws_analysis)
    
    for type in raw_material_types:
        total_needed = sum(forecast[date][type.id] for date in forecast) if forecast else 0
        current = current_stock.get(type.id, 0)
        balance = current - total_needed
        status = "Достаточно" if balance >= 0 else f"Требуется докупить {abs(balance):.2f} кг"
        
        ws_analysis.append([
            type.name,
            current,
            total_needed,
            balance,
            status
        ])
    
    adjust_column_width(ws_analysis)
    
    # Сохраняем и отправляем файл
    output = save_excel_report(wb, output)
    filename = f"raw_material_forecast_{format_datetime(datetime.now())}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/reports/production_plans/export')
@login_required
def export_production_plans():
    """Экспорт отчёта по планам производства в Excel"""
    wb, output = create_excel_report()
    
    # Создаем лист для отчета
    ws = wb.create_sheet("Планы производства")
    
    # Заголовки
    headers = ["Дата внесения в план", "Продукт", "Партия", "Количество (кг)", "Статус", "Номер недели", "Отслеживание", "Контроль ОКК"]
    ws.append(headers)
    style_header_row(ws)
    
    # Сопоставление статусов для вывода на русском
    status_map = {
        "черновик": "Черновик",
        "утверждён": "Утверждён",
        "в производстве": "В производстве",
        "завершен": "Завершен",
        "отменен": "Отменен"
    }
    
    # Получаем данные
    plans = ProductionPlan.query.order_by(ProductionPlan.created_at.desc()).all()
    
    for plan in plans:
        # Вычисляем прогресс выполнения
        total_produced = sum(batch.weight for batch in plan.batches)
        progress_percent = (total_produced / plan.quantity * 100) if plan.quantity > 0 else 0
        
        # Получаем человекочитаемый статус
        if hasattr(plan.status, "display"):
            status_display = plan.status.display
        elif hasattr(plan.status, "value"):
            status_display = status_map.get(plan.status.value, plan.status.value)
        else:
            status_display = status_map.get(str(plan.status), str(plan.status))
        
        # Вычисляем номер недели
        week_number = plan.created_at.isocalendar()[1]
        
        ws.append([
            plan.created_at.strftime("%Y-%m-%d"),
            plan.product.name,
            plan.batch_number,
            plan.quantity,
            status_display,
            week_number,
            "",  # Пустая ячейка для "Отслеживание"
            f"{progress_percent:.1f}"
        ])
    
    # Форматирование
    adjust_column_width(ws)
    
    # Сохраняем и отправляем файл
    output = save_excel_report(wb, output)
    filename = f"production_plans_{format_datetime(datetime.now())}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/production_plans/<int:plan_id>/export_used_materials')
@login_required
def export_used_materials(plan_id):
    """Экспорт отчёта по использованному сырью для завершённого плана"""
    plan = ProductionPlan.query.get_or_404(plan_id)
    if plan.status != PlanStatus.COMPLETED:
        flash('Отчёт доступен только для завершённых планов!', 'error')
        return redirect(url_for('production_plan_detail', plan_id=plan_id))

    # Сбор данных: (type_name, batch_number) -> total_quantity
    usage = {}
    for batch in plan.batches:
        for bm in batch.materials:
            type_name = bm.material_batch.material.type.name
            batch_number = bm.material_batch.material.batch_number
            key = (type_name, batch_number)
            usage.setdefault(key, 0)
            usage[key] += bm.quantity

    # Формируем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Использованное сырьё"
    ws.append(["Вид сырья", "Партия сырья", "Количество (кг)"])
    for (type_name, batch_number), qty in usage.items():
        ws.append([type_name, batch_number, round(qty, 3)])
    # Форматирование ширины
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    # Сохраняем в память
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    # Получаем имя продукта и номер партии плана
    product_name = plan.product.name.replace(' ', '_')
    batch_number = plan.batch_number
    filename = f"{product_name}_{batch_number}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/production_plans/<int:plan_id>/add_multiple_batches', methods=['POST'])
@operator_required
def add_multiple_batches(plan_id):
    plan = ProductionPlan.query.get_or_404(plan_id)
    
    try:
        # Получаем данные из формы
        num_batches = int(request.form.get('num_batches', 0))
        weight_per_batch = float(request.form.get('weight_per_batch', 0))
        
        if num_batches <= 0 or weight_per_batch <= 0:
            flash('Количество замесов и вес должны быть больше 0', 'error')
            return redirect(url_for('production_plan_detail', plan_id=plan_id))
        
        # Проверяем максимальный вес замеса
        if weight_per_batch > 1000:
            flash('Вес замеса не может превышать 1000 кг', 'error')
            return redirect(url_for('production_plan_detail', plan_id=plan_id))
        
        # Проверяем, не превышает ли общее количество план
        total_quantity = sum([batch.weight for batch in plan.batches]) + (num_batches * weight_per_batch)
        if total_quantity > plan.quantity:
            flash(f'Общее количество замесов ({total_quantity} кг) превышает план ({plan.quantity} кг)', 'error')
            return redirect(url_for('production_plan_detail', plan_id=plan_id))
        
        # Находим следующий номер замеса
        existing_batch_numbers = [batch.batch_number for batch in plan.batches]
        next_batch_number = 1
        while str(next_batch_number) in existing_batch_numbers:
            next_batch_number += 1
        
        created_batches = []
        failed_batches = []
        
        # Создаём замесы
        for i in range(num_batches):
            batch_number = str(next_batch_number + i)
            
            batch = ProductionBatch(
                plan=plan,
                batch_number=batch_number,
                weight=weight_per_batch
            )
            db.session.add(batch)
            db.session.flush()  # Получаем ID замеса
            
            # Автоматически добавляем все ингредиенты из рецептуры
            batch_ingredients = []
            missing_ingredients = []
            
            for recipe_item in plan.template.recipe_items:
                needed_qty = batch.weight * float(recipe_item.percentage) / 100
                
                # Получаем список партий с нужным количеством
                materials_to_use, shortage = get_available_materials_for_batch(
                    needed_qty, recipe_item.material_type_id
                )
                
                if shortage > 0:
                    missing_ingredients.append(f"{recipe_item.material_type.name} (нужно {needed_qty:.2f} кг, недостаточно {shortage:.2f} кг)")
                    continue
                
                # Создаём записи для каждой использованной партии
                for material_info in materials_to_use:
                    material = material_info['material']
                    qty = material_info['quantity']
                    
                    # Создаём MaterialBatch и BatchMaterial
                    material_batch = MaterialBatch(
                        material=material,
                        batch_number=material.batch_number,
                        quantity=qty,
                        remaining_quantity=qty
                    )
                    db.session.add(material_batch)
                    db.session.flush()
                    
                    ingredient = BatchMaterial(
                        batch=batch,
                        material_batch=material_batch,
                        quantity=qty
                    )
                    db.session.add(ingredient)
                
                # Добавляем информацию об ингредиенте
                total_used = sum(m['quantity'] for m in materials_to_use)
                if len(materials_to_use) == 1:
                    batch_ingredients.append(f"{recipe_item.material_type.name} ({total_used:.2f} кг из партии {materials_to_use[0]['material'].batch_number})")
                else:
                    batches_info = [f"{m['material'].batch_number}({m['quantity']:.2f} кг)" for m in materials_to_use]
                    batch_ingredients.append(f"{recipe_item.material_type.name} ({total_used:.2f} кг из партий: {', '.join(batches_info)})")
            
            if missing_ingredients:
                failed_batches.append(f"Замес №{batch_number}: недостаточно сырья - {', '.join(missing_ingredients)}")
                db.session.rollback()
                continue
            else:
                created_batches.append(f"Замес №{batch_number} ({weight_per_batch} кг): {', '.join(batch_ingredients)}")
        
        # Добавляем запись в примечания
        timestamp = datetime.now().strftime('%d.%m.%Y %H:%M')
        batch_note = f"[{timestamp}] Добавлено {len(created_batches)} замесов по {weight_per_batch} кг каждый"
        
        if plan.notes:
            plan.notes = batch_note + "\n\n" + plan.notes
        else:
            plan.notes = batch_note
            
        db.session.commit()
        
        if created_batches:
            flash(f'Успешно создано {len(created_batches)} замесов', 'success')
        
        if failed_batches:
            flash(f'Не удалось создать {len(failed_batches)} замесов из-за недостатка сырья', 'warning')
            
    except ValueError:
        flash('Некорректные данные в форме', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при создании замесов: {str(e)}', 'error')
    
    return redirect(url_for('production_plan_detail', plan_id=plan_id))

@app.route('/production_plans/<int:plan_id>/delete_all_batches', methods=['POST'])
@operator_required
def delete_all_batches(plan_id):
    plan = ProductionPlan.query.get_or_404(plan_id)
    
    # Проверяем, что план не завершён
    if plan.status == PlanStatus.COMPLETED:
        flash('Нельзя удалять замесы из завершённого плана', 'error')
        return redirect(url_for('production_plan_detail', plan_id=plan_id))
    
    # Подсчитываем количество удаляемых замесов
    num_batches = len(plan.batches)
    
    if num_batches == 0:
        flash('В плане нет замесов для удаления', 'info')
        return redirect(url_for('production_plan_detail', plan_id=plan_id))
    
    try:
        # Удаляем все замесы (cascade="all, delete-orphan" в модели автоматически удалит связанные записи)
        for batch in plan.batches:
            db.session.delete(batch)
        
        # Добавляем запись в примечания
        timestamp = datetime.now().strftime('%d.%m.%Y %H:%M')
        batch_note = f"[{timestamp}] Удалено {num_batches} замесов"
        
        if plan.notes:
            plan.notes = batch_note + "\n\n" + plan.notes
        else:
            plan.notes = batch_note
            
        db.session.commit()
        flash(f'Удалено {num_batches} замесов', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при удалении замесов: {str(e)}', 'error')
    
    return redirect(url_for('production_plan_detail', plan_id=plan_id))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = LoginForm()
    if form.validate_on_submit():
        from app.models import User
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            login_user(user)
            flash('Вы успешно вошли в систему!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            flash('Неверное имя пользователя или пароль', 'error')
    return render_template('login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Вы вышли из системы.', 'info')
    return redirect(url_for('login'))

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(e):
    return render_template('500.html'), 500

@app.route('/health')
def health_check():
    from datetime import datetime
    return {
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'service': 'Production Planner',
        'version': '1.0.0'
    }

@app.route('/apply_migrations')
@admin_required
def apply_migrations():
    """Применяет миграции базы данных"""
    try:
        result = subprocess.run(
            ['alembic', '-c', 'migrations/alembic.ini', 'upgrade', 'head'],
            capture_output=True,
            text=True,
            check=True
        )
        return f"Миграции применены успешно!<br><pre>{result.stdout}</pre>"
    except subprocess.CalledProcessError as e:
        return f"Ошибка при применении миграций:<br><pre>{e.stderr}</pre>"
    except FileNotFoundError:
        return "Alembic не найден" 

@app.route('/production_plans/<int:plan_id>/edit_quantity', methods=['POST'])
@operator_required
def edit_plan_quantity(plan_id):
    plan = ProductionPlan.query.get_or_404(plan_id)
    
    # Проверяем, что план имеет статус "Черновик"
    if plan.status != PlanStatus.DRAFT:
        flash('Можно изменять количество только в планах со статусом "Черновик"', 'error')
        return redirect(url_for('production_plan_detail', plan_id=plan_id))
    
    try:
        # Получаем новое количество из формы
        new_quantity = float(request.form.get('new_quantity', 0))
        
        if new_quantity <= 0:
            flash('Количество должно быть больше 0', 'error')
            return redirect(url_for('production_plan_detail', plan_id=plan_id))
        
        # Сохраняем старое количество для записи в примечания
        old_quantity = plan.quantity
        
        # Обновляем количество в плане
        plan.quantity = new_quantity
        
        # Добавляем запись в примечания
        timestamp = datetime.now().strftime('%d.%m.%Y %H:%M')
        quantity_note = f"[{timestamp}] Изменено количество с {old_quantity:.2f} кг на {new_quantity:.2f} кг"
        
        if plan.notes:
            plan.notes = quantity_note + "\n\n" + plan.notes
        else:
            plan.notes = quantity_note
            
        db.session.commit()
        flash(f'Количество плана изменено с {old_quantity:.2f} кг на {new_quantity:.2f} кг', 'success')
        
    except ValueError:
        flash('Некорректное значение количества', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при изменении количества: {str(e)}', 'error')
    
    return redirect(url_for('production_plan_detail', plan_id=plan_id))

@app.route('/production_plans/<int:plan_id>/edit_name', methods=['POST'])
@operator_required
def edit_plan_name(plan_id):
    plan = ProductionPlan.query.get_or_404(plan_id)
    
    # Проверяем, что план имеет статус "Черновик"
    if plan.status != PlanStatus.DRAFT:
        flash('Можно изменять номер партии только в планах со статусом "Черновик"', 'error')
        return redirect(url_for('production_plan_detail', plan_id=plan_id))
    
    try:
        # Получаем новый номер партии из формы
        new_batch_number = request.form.get('new_name', '').strip()
        
        if not new_batch_number:
            flash('Номер партии не может быть пустым', 'error')
            return redirect(url_for('production_plan_detail', plan_id=plan_id))
        
        # Сохраняем старый номер партии для записи в примечания
        old_batch_number = plan.batch_number
        
        # Обновляем номер партии в плане
        plan.batch_number = new_batch_number
        
        # Добавляем запись в примечания
        timestamp = datetime.now().strftime('%d.%m.%Y %H:%M')
        batch_number_note = f"[{timestamp}] Изменен номер партии с '{old_batch_number}' на '{new_batch_number}'"
        
        if plan.notes:
            plan.notes = batch_number_note + "\n\n" + plan.notes
        else:
            plan.notes = batch_number_note
            
        db.session.commit()
        flash(f'Номер партии изменен с "{old_batch_number}" на "{new_batch_number}"', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при изменении номера партии: {str(e)}', 'error')
    
    return redirect(url_for('production_plan_detail', plan_id=plan_id))

def restore_raw_materials_from_plan(plan):
    """Автоматически восстанавливает сырьё из завершённого плана"""
    
    restored_count = 0
    restored_materials = []
    
    try:
        for batch in plan.batches:
            for batch_material in batch.materials:
                # Находим соответствующее сырьё
                raw_material = batch_material.material_batch.material
                
                if raw_material:
                    # Восстанавливаем количество
                    raw_material.quantity_kg += batch_material.quantity
                    restored_count += 1
                    
                    restored_materials.append(f"{raw_material.batch_number}: +{batch_material.quantity:.2f} кг")
        
        return f"{restored_count} позиций: {', '.join(restored_materials)}"
        
    except Exception as e:
        print(f"Ошибка при восстановлении сырья: {e}")
        return f"Ошибка: {e}"

@app.route('/production_plans/<int:plan_id>/undo_completion', methods=['POST'])
@admin_required
def undo_plan_completion(plan_id):
    """Отменяет завершение плана с автоматическим восстановлением сырья"""
    
    plan = ProductionPlan.query.get_or_404(plan_id)
    
    if plan.status != PlanStatus.COMPLETED:
        flash('Можно отменить только завершённые планы', 'error')
        return redirect(url_for('production_plan_detail', plan_id=plan_id))
    
    try:
        # 1. Автоматически восстанавливаем сырьё
        restored_materials = restore_raw_materials_from_plan(plan)
        
        # 2. Меняем статус на "Черновик"
        plan.status = PlanStatus.DRAFT
        
        # 3. Добавляем запись в notes
        timestamp = datetime.now().strftime('%d.%m.%Y %H:%M')
        undo_note = f"[{timestamp}] Отменено завершение плана. Восстановлено сырья: {restored_materials}"
        
        if plan.notes:
            plan.notes = undo_note + "\n\n" + plan.notes
        else:
            plan.notes = undo_note
        
        db.session.commit()
        
        flash(f'Завершение плана отменено! Восстановлено сырья: {restored_materials}', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при отмене завершения: {e}', 'error')
    
    return redirect(url_for('production_plan_detail', plan_id=plan_id))

@app.route('/products/<int:product_id>/edit_recipe', methods=['GET', 'POST'])
@admin_required
def edit_product_recipe(product_id):
    product = Product.query.get_or_404(product_id)
    
    if request.method == 'GET':
        # Получаем текущую рецептуру
        if not product.recipe_templates:
            flash('У продукта нет рецептуры для редактирования', 'error')
            return redirect(url_for('products'))
        
        # Берём первую рецептуру (или можно добавить выбор конкретной)
        recipe_template = product.recipe_templates[0]
        recipe_items = RecipeItem.query.filter_by(template_id=recipe_template.id).order_by(RecipeItem.id).all()
        raw_material_types = RawMaterialType.query.order_by(RawMaterialType.name).all()
        
        return render_template(
            'edit_product_recipe.html',
            product=product,
            recipe_template=recipe_template,
            recipe_items=recipe_items,
            raw_material_types=raw_material_types
        )
    
    elif request.method == 'POST':
        try:
            # Проверяем наличие рецептуры
            if not product.recipe_templates:
                flash('У продукта нет рецептуры для редактирования', 'error')
                return redirect(url_for('products'))
            
            recipe_template = product.recipe_templates[0]
            
            # Получаем данные из формы
            material_type_ids = request.form.getlist('material_type_id[]')
            percentages = request.form.getlist('percentage[]')
            
            # Валидация данных
            if not material_type_ids or not percentages:
                flash('Необходимо указать хотя бы один ингредиент', 'error')
                return redirect(url_for('edit_product_recipe', product_id=product_id))
            
            if len(material_type_ids) != len(percentages):
                flash('Ошибка в данных формы', 'error')
                return redirect(url_for('edit_product_recipe', product_id=product_id))
            
            # Проверяем, что все проценты положительные
            total_percentage = 0
            for percentage_str in percentages:
                try:
                    percentage = float(percentage_str)
                    if percentage <= 0:
                        flash('Процент каждого ингредиента должен быть больше 0', 'error')
                        return redirect(url_for('edit_product_recipe', product_id=product_id))
                    total_percentage += percentage
                except ValueError:
                    flash('Некорректное значение процента', 'error')
                    return redirect(url_for('edit_product_recipe', product_id=product_id))
            
            # Проверяем, что сумма процентов = 100%
            if abs(total_percentage - 100.0) > 0.001:
                flash(f'Сумма процентов должна быть равна 100%. Текущая сумма: {total_percentage:.3f}%', 'error')
                return redirect(url_for('edit_product_recipe', product_id=product_id))
            
            # Проверяем уникальность типов сырья
            if len(set(material_type_ids)) != len(material_type_ids):
                flash('Каждый тип сырья может быть указан только один раз', 'error')
                return redirect(url_for('edit_product_recipe', product_id=product_id))
            
            # Проверяем, не используется ли продукт в активных планах
            active_plans = ProductionPlan.query.filter(
                ProductionPlan.product_id == product_id,
                ProductionPlan.status.in_([PlanStatus.APPROVED, PlanStatus.IN_PROGRESS])
            ).first()
            
            if active_plans:
                flash('Нельзя изменять рецептуру продукта, который используется в активных планах производства', 'error')
                return redirect(url_for('edit_product_recipe', product_id=product_id))
            
            # Начинаем транзакцию
            db.session.begin()
            
            # Удаляем старые записи рецептуры
            RecipeItem.query.filter_by(template_id=recipe_template.id).delete()
            
            # Создаём новые записи рецептуры
            for material_type_id, percentage_str in zip(material_type_ids, percentages):
                percentage = float(percentage_str)
                recipe_item = RecipeItem(
                    template_id=recipe_template.id,
                    material_type_id=int(material_type_id),
                    percentage=percentage
                )
                db.session.add(recipe_item)
            
            # Сохраняем изменения
            db.session.commit()
            
            flash('Рецептура продукта успешно обновлена', 'success')
            return redirect(url_for('products'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при обновлении рецептуры: {str(e)}', 'error')
            return redirect(url_for('edit_product_recipe', product_id=product_id))

@app.route('/admin/cleanup_orphaned_data')
@admin_required
def cleanup_orphaned_data():
    """Временный маршрут для очистки 'висящих' ссылок на удалённое сырьё"""
    
    try:
        # Инициализируем переменные SQL диагностики
        orphaned_mb_sql = []
        orphaned_bm_sql = []
        orphaned_pb_sql = []
        
        # Диагностика: находим "висящие" ссылки
        orphaned_material_batches = db.session.query(MaterialBatch).filter(
            ~MaterialBatch.material_id.in_(
                db.session.query(RawMaterial.id)
            )
        ).all()
        
        orphaned_batch_materials = db.session.query(BatchMaterial).filter(
            ~BatchMaterial.material_batch_id.in_(
                db.session.query(MaterialBatch.id)
            )
        ).all()
        
        # Дополнительная диагностика: находим "висящие" ссылки на уровне SQL
        try:
            from sqlalchemy import text
            
            # Проверяем MaterialBatch с несуществующим material_id
            orphaned_mb_sql = db.session.execute(text("""
                SELECT mb.id, mb.material_id, mb.batch_number, mb.quantity
                FROM material_batches mb
                LEFT JOIN raw_materials rm ON mb.material_id = rm.id
                WHERE rm.id IS NULL
            """)).fetchall()
            
            # Проверяем BatchMaterial с несуществующим material_batch_id
            orphaned_bm_sql = db.session.execute(text("""
                SELECT bm.id, bm.material_batch_id, bm.batch_id, bm.quantity
                FROM batch_materials bm
                LEFT JOIN material_batches mb ON bm.material_batch_id = mb.id
                WHERE mb.id IS NULL
            """)).fetchall()
            
            # Проверяем ProductionBatch с несуществующим plan_id
            orphaned_pb_sql = db.session.execute(text("""
                SELECT pb.id, pb.plan_id, pb.batch_number, pb.weight
                FROM production_batches pb
                LEFT JOIN production_plans pp ON pb.plan_id = pp.id
                WHERE pp.id IS NULL
            """)).fetchall()
            
        except Exception as e:
            orphaned_mb_sql = []
            orphaned_bm_sql = []
            orphaned_pb_sql = []
            print(f"Ошибка SQL диагностики: {e}")
        
        # Диагностика: находим планы с проблемами
        problematic_plans = []
        for plan in ProductionPlan.query.all():
            plan_problems = []
            
            try:
                # Проверяем продукт
                if not plan.product:
                    plan_problems.append("Продукт не найден")
                
                # Проверяем рецептуру
                if not plan.template:
                    plan_problems.append("Рецептура не найдена")
                
                # Проверяем замесы
                for batch in plan.batches:
                    batch_problems = []
                    
                    # Проверяем ингредиенты замеса
                    for batch_material in batch.materials:
                        ingredient_problems = []
                        
                        # Проверяем MaterialBatch
                        if not batch_material.material_batch:
                            ingredient_problems.append("MaterialBatch не найден")
                        else:
                            # Проверяем сырьё
                            if not batch_material.material_batch.material:
                                ingredient_problems.append("Сырьё не найдено")
                            else:
                                # Проверяем тип сырья
                                if not batch_material.material_batch.material.type:
                                    ingredient_problems.append("Тип сырья не найден")
                        
                        if ingredient_problems:
                            batch_problems.append(f"Ингредиент {batch_material.id}: {', '.join(ingredient_problems)}")
                    
                    if batch_problems:
                        plan_problems.append(f"Замес {batch.batch_number}: {', '.join(batch_problems)}")
                
                # Если есть проблемы, добавляем план в список
                if plan_problems:
                    problematic_plans.append({
                        'id': plan.id,
                        'batch_number': plan.batch_number,
                        'product': plan.product.name if plan.product else 'Неизвестно',
                        'status': plan.status,
                        'problems': plan_problems,
                        'problem_count': len(plan_problems)
                    })
                    
            except Exception as e:
                problematic_plans.append({
                    'id': plan.id,
                    'batch_number': plan.batch_number,
                    'product': 'Ошибка загрузки',
                    'status': 'Ошибка',
                    'problems': [f'Критическая ошибка: {str(e)}'],
                    'problem_count': 1
                })
        
        if request.method == 'POST':
            # Выполняем очистку
            action = request.form.get('action')
            
            if action == 'cleanup_material_batches':
                # Удаляем "висящие" MaterialBatch
                count = 0
                for orphan in orphaned_material_batches:
                    db.session.delete(orphan)
                    count += 1
                db.session.commit()
                flash(f'Удалено {count} "висящих" MaterialBatch', 'success')
                
            elif action == 'cleanup_batch_materials':
                # Удаляем "висящие" BatchMaterial
                count = 0
                for orphan in orphaned_batch_materials:
                    db.session.delete(orphan)
                    count += 1
                db.session.commit()
                flash(f'Удалено {count} "висящих" BatchMaterial', 'success')
                
            elif action == 'cleanup_all':
                # Удаляем всё "висящее"
                count_mb = 0
                for orphan in orphaned_material_batches:
                    db.session.delete(orphan)
                    count_mb += 1
                
                count_bm = 0
                for orphan in orphaned_batch_materials:
                    db.session.delete(orphan)
                    count_bm += 1
                
                db.session.commit()
                flash(f'Удалено {count_mb} MaterialBatch и {count_bm} BatchMaterial', 'success')
            
            # Перезагружаем страницу для обновления данных
            return redirect(url_for('cleanup_orphaned_data'))
        
        return render_template(
            'cleanup_orphaned_data.html',
            orphaned_material_batches=orphaned_material_batches,
            orphaned_batch_materials=orphaned_batch_materials,
            problematic_plans=problematic_plans,
            orphaned_mb_sql=orphaned_mb_sql,
            orphaned_bm_sql=orphaned_bm_sql,
            orphaned_pb_sql=orphaned_pb_sql
        )
        
    except Exception as e:
        flash(f'Ошибка при диагностике: {str(e)}', 'error')
        return redirect(url_for('index'))