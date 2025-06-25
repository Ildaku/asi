from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io

def style_header_row(ws, row=1):
    """Стилизация заголовков таблицы"""
    header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

def adjust_column_width(ws):
    """Автоматическая настройка ширины столбцов"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_excel_report():
    """Создание нового Excel файла с базовыми настройками"""
    wb = Workbook()
    wb.remove(wb.active)  # Удаляем стандартный лист
    output = io.BytesIO()
    return wb, output

def save_excel_report(wb, output):
    """Сохранение Excel файла в BytesIO"""
    wb.save(output)
    output.seek(0)
    return output

def format_datetime(dt):
    """Форматирование даты и времени для имени файла"""
    return dt.strftime("%Y%m%d_%H%M%S") 